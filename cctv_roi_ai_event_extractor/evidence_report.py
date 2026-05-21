import os
import re
from copy import copy
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VIDEO_DATETIME_PATTERNS = (
    re.compile(r"(?P<date>20\d{6})[_\-\u2013\u2014](?P<time>\d{6})"),
    re.compile(r"p(?P<yy>\d{2})(?P<mm>\d{2})(?P<dd>\d{2})[_\-\u2013\u2014](?P<time>\d{6})", re.IGNORECASE),
)


def parse_video_start_datetime(video_rel_path: str):
    """Parse start time from known CCTV filename formats."""
    name = os.path.basename(video_rel_path or "")
    for pattern in VIDEO_DATETIME_PATTERNS:
        match = pattern.search(name)
        if not match:
            continue
        groups = match.groupdict()
        if "date" in groups and groups.get("date"):
            raw_date = groups["date"]
            raw_time = groups["time"]
            return datetime.strptime(raw_date + raw_time, "%Y%m%d%H%M%S")
        year = 2000 + int(groups["yy"])
        raw = f"{year:04d}{groups['mm']}{groups['dd']}{groups['time']}"
        return datetime.strptime(raw, "%Y%m%d%H%M%S")
    return None


def format_event_datetime(video_rel_path: str, seconds_text: str):
    """Convert a video-relative event timestamp to an absolute timestamp when possible."""
    if seconds_text is None or str(seconds_text).strip() == "":
        return ""
    start = parse_video_start_datetime(video_rel_path)
    if start is None:
        return ""
    try:
        seconds = float(seconds_text or 0)
    except (TypeError, ValueError):
        seconds = 0.0
    return (start + timedelta(seconds=seconds)).strftime("%Y-%m-%d %H:%M:%S")


def _first_existing_path(paths):
    for path in paths:
        if path and os.path.exists(path):
            return path
    return None


def _parse_seconds(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _parse_bbox(value):
    try:
        parts = [int(float(part.strip())) for part in str(value or "").split(",")]
    except (TypeError, ValueError):
        return None
    if len(parts) != 4:
        return None
    return tuple(parts)


def _bbox_iou(a, b):
    if not a or not b:
        return 0.0
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)
    inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
    if inter <= 0:
        return 0.0
    area_a = max(0, ax2 - ax1) * max(0, ay2 - ay1)
    area_b = max(0, bx2 - bx1) * max(0, by2 - by1)
    union = area_a + area_b - inter
    return inter / float(union) if union > 0 else 0.0


def _plate_row_score(item):
    """Score LPR rows so valid, confident, plate-length text wins deduplication."""
    try:
        confidence = max(float(part) for part in str(item.get("plate_confidence", "0")).split(";") if part)
    except ValueError:
        confidence = 0.0
    valid_bonus = 1.0 if "Y" in str(item.get("plate_valid_taiwan_format", "")) else 0.0
    text = str(item.get("plate_text", ""))
    length_bonus = 0.2 if 4 <= len(text) <= 8 else -1.0
    return valid_bonus + confidence + length_bonus


def _select_best_lpr_row(group):
    """Pick the strongest row after grouping repeated sightings of the same plate."""
    text_scores = {}
    for item in group:
        text = str(item.get("plate_text", ""))
        text_scores[text] = text_scores.get(text, 0.0) + 2.0 + _plate_row_score(item)
    best_text = max(text_scores, key=text_scores.get)
    best_rows = [item for item in group if str(item.get("plate_text", "")) == best_text]
    return max(best_rows, key=_plate_row_score)


def _dedupe_lpr_rows(rows):
    """Merge repeated LPR detections by video, time proximity, and plate-box overlap."""
    groups = []
    for row in sorted(rows, key=lambda item: _parse_seconds(item.get("event_time_sec"))):
        row_bbox = _parse_bbox(row.get("plate_bbox"))
        row_time = _parse_seconds(row.get("event_time_sec"))
        matched_index = None
        for idx, group in enumerate(groups):
            if row.get("video_rel_path") != group[0].get("video_rel_path"):
                continue
            group_times = [_parse_seconds(existing.get("event_time_sec")) for existing in group]
            if min(abs(row_time - existing_time) for existing_time in group_times) > 60.0:
                continue
            if any(_bbox_iou(row_bbox, _parse_bbox(existing.get("plate_bbox"))) >= 0.10 for existing in group):
                matched_index = idx
                break
        if matched_index is None:
            groups.append([row])
        else:
            groups[matched_index].append(row)
    return [_select_best_lpr_row(group) for group in groups]


def _track_row_identity(item):
    camera_id = str(item.get("camera_id", ""))
    track_id = str(item.get("track_id", ""))
    if camera_id or track_id:
        return camera_id, track_id
    return str(item.get("track_start_source", "")), str(item.get("track_id", ""))


def _dedupe_track_rows(rows):
    """Keep one best row per tracked vehicle identity."""
    groups = {}
    for item in rows:
        key = _track_row_identity(item)
        current = groups.get(key)
        if current is None or _plate_row_score(item) > _plate_row_score(current):
            groups[key] = item
    return list(groups.values())


def build_evidence_rows(csv_rows):
    """Convert raw processing log rows into workbook-ready evidence rows."""
    rows = []
    track_rows = _dedupe_track_rows([item for item in csv_rows if item.get("record_type") == "track_summary"])
    for item in track_rows:
        plate_text = item.get("plate_text", "")
        if not plate_text:
            continue
        rows.append({
            "entry_datetime": item.get("track_start_datetime") or format_event_datetime(item.get("video_rel_path", ""), item.get("event_time_sec", "")),
            "exit_datetime": item.get("track_end_datetime") or format_event_datetime(item.get("video_rel_path", ""), item.get("interval_end_sec", "")),
            "plate_text": plate_text,
            "screenshot_path": _first_existing_path([item.get("output_path")]) or "",
        })
    if rows:
        return rows

    lpr_rows = [item for item in csv_rows if item.get("record_type") == "lpr_detection" and item.get("plate_text")]
    screenshot_rows = [item for item in csv_rows if item.get("record_type") == "screenshot"]
    for item in _dedupe_lpr_rows(lpr_rows) + screenshot_rows:
        record_type = item.get("record_type")
        if record_type not in {"screenshot", "lpr_detection"}:
            continue
        plate_text = item.get("plate_text", "")
        if record_type == "screenshot" and not plate_text:
            continue
        screenshot_path = _first_existing_path([item.get("output_path")])
        if record_type == "screenshot" and not screenshot_path:
            continue
        rows.append({
            "entry_datetime": format_event_datetime(item.get("video_rel_path", ""), item.get("event_time_sec", "")),
            "exit_datetime": format_event_datetime(item.get("video_rel_path", ""), item.get("interval_end_sec", "")),
            "plate_text": plate_text,
            "screenshot_path": screenshot_path or "",
        })
    return rows


def write_evidence_workbook(output_path: str, csv_rows):
    """Write the evidence workbook and embed screenshots when image files exist."""
    rows = build_evidence_rows(csv_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "蒐證資料"

    headers = ["編號", "進入日期", "出去日期", "車號", "車輛截圖"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 34

    for idx, item in enumerate(rows, start=1):
        row_idx = idx + 1
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=item["entry_datetime"])
        ws.cell(row=row_idx, column=3, value=item["exit_datetime"])
        ws.cell(row=row_idx, column=4, value=item["plate_text"])
        ws.row_dimensions[row_idx].height = 82

        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

        if not item["screenshot_path"]:
            continue
        try:
            img = ExcelImage(item["screenshot_path"])
            img.width = 160
            img.height = 90
            ws.add_image(img, f"E{row_idx}")
        except Exception:
            ws.cell(row=row_idx, column=5, value=item["screenshot_path"])

    for row in ws.iter_rows(min_row=1, max_row=max(1, len(rows) + 1), min_col=1, max_col=5):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "center"
            cell.alignment = alignment

    ensure_dir = os.path.dirname(output_path)
    if ensure_dir:
        os.makedirs(ensure_dir, exist_ok=True)
    wb.save(output_path)
    return output_path, len(rows)
